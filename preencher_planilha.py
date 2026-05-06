import sys
import re
import glob
import calendar
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# ⚙️ Configure antes de executar
VALOR_FIXO_AERO = 2.07
ARQUIVO_OFICIAL = 'INFORMATIVOS_FALECIDOS.2026.xlsx'

FORMATO_MOEDA = r'_-"R$ "* #,##0.00_-;\-"R$ "* #,##0.00_-;_-"R$ "* "-"??_-;_-@_-'
MESES_ABREV = {
    1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun',
    7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez',
}


def clean_cpf(cpf) -> str:
    if pd.isna(cpf):
        return ''
    if isinstance(cpf, float):
        cpf = str(int(cpf))
    return re.sub(r'\D', '', str(cpf)).zfill(11)


def proximo_mes(dt: pd.Timestamp) -> pd.Timestamp:
    if dt.month == 12:
        return dt.replace(year=dt.year + 1, month=1)
    return dt.replace(month=dt.month + 1)


def formatar_periodo(meses: list) -> str:
    if not meses:
        return '-'
    def fmt(dt):
        return f"{MESES_ABREV[dt.month]}/{str(dt.year)[-2:]}"
    inicio = fmt(meses[0])
    fim = fmt(meses[-1])
    return inicio if inicio == fim else f"{inicio} a {fim}"


def carregar_demonstrativos() -> pd.DataFrame:
    xls = pd.ExcelFile('DEMONSTRATIVOS (POR ANO).xlsx')
    abas = [a for a in xls.sheet_names if str(a).isdigit() and len(str(a)) == 4]
    df = pd.concat([pd.read_excel(xls, sheet_name=a) for a in abas], ignore_index=True)
    df['CPF_CLEAN'] = df['CPF BENEFICIÁRIO'].apply(clean_cpf)
    df['COMPETÊNCIA'] = pd.to_datetime(df['COMPETÊNCIA'], errors='coerce')
    df = df[df['TIPO LANÇAMENTO'].str.upper().str.contains('MENSALIDADE', na=False)]
    df['CLASSIFICAÇÃO'] = df['CLASSIFICAÇÃO'].str.upper().str.replace('Ú', 'U')

    grp = df.groupby(
        ['CPF_CLEAN', 'COMPETÊNCIA', 'CLASSIFICAÇÃO'], as_index=False
    )[['VALOR OPERADORA', 'SUBSDIO']].sum()

    pivot = grp.pivot(
        index=['CPF_CLEAN', 'COMPETÊNCIA'],
        columns='CLASSIFICAÇÃO',
        values=['VALOR OPERADORA', 'SUBSDIO'],
    ).reset_index()
    pivot.columns = [f"{c[0]}_{c[1]}" if c[1] else c[0] for c in pivot.columns]
    pivot.rename(columns={
        'VALOR OPERADORA_SAUDE': 'MENSALIDADE_SAUDE',
        'SUBSDIO_SAUDE': 'SUBSIDIO_SAUDE',
        'VALOR OPERADORA_ODONTO': 'MENSALIDADE_ODONTO',
        'SUBSDIO_ODONTO': 'SUBSIDIO_ODONTO',
    }, inplace=True)
    return pivot


def calcular_rateio(row, df_dem_pivot: pd.DataFrame) -> pd.Series:
    cpf = row['CPF_CLEAN']
    dt_obito = row['DATA_OBITO']
    dt_exclusao = row['DATA_CANCELAMENTO']
    tem_aero = str(row.get('AERO', 'NÃO')).strip().upper() == 'SIM'

    vazio = pd.Series({
        'ESPERADO_SAUDE': 0.0,
        'PCT_PBH_SAUDE': 0.0,
        'PCT_PBH_ODONTO': 0.0,
        'PERIODO_SUBSIDIADO_STR': '-',
    })

    if pd.isna(dt_obito) or pd.isna(dt_exclusao) or dt_exclusao <= dt_obito:
        return vazio

    mes_obito = dt_obito.replace(day=1)
    mes_fim = dt_exclusao.replace(day=1)
    df_cpf = df_dem_pivot[df_dem_pivot['CPF_CLEAN'] == cpf]
    dem_obito = df_cpf[df_cpf['COMPETÊNCIA'] == mes_obito]

    esp_saude = esp_odonto = 0.0
    pbh_saude = pbh_odonto = 0.0
    meses_subsidiados = []
    mes_atual = mes_obito

    while mes_atual <= mes_fim:
        dem_mes = df_cpf[df_cpf['COMPETÊNCIA'] == mes_atual]

        if not dem_mes.empty:
            m_saude = float(dem_mes['MENSALIDADE_SAUDE'].iloc[0] or 0)
            s_saude = float(dem_mes['SUBSIDIO_SAUDE'].iloc[0] or 0)
            m_odonto = float(dem_mes['MENSALIDADE_ODONTO'].iloc[0] or 0)
            s_odonto = float(dem_mes['SUBSIDIO_ODONTO'].iloc[0] or 0)
        else:
            m_saude = float(dem_obito['MENSALIDADE_SAUDE'].iloc[0] or 0) if not dem_obito.empty else 0.0
            m_odonto = float(dem_obito['MENSALIDADE_ODONTO'].iloc[0] or 0) if not dem_obito.empty else 0.0
            s_saude = s_odonto = 0.0

        if s_saude > 0 or s_odonto > 0:
            meses_subsidiados.append(mes_atual)

        if tem_aero:
            m_saude += VALOR_FIXO_AERO

        dias_no_mes = calendar.monthrange(mes_atual.year, mes_atual.month)[1]
        dias_a_devolver = (dias_no_mes - dt_obito.day) if mes_atual == mes_obito else dias_no_mes
        fracao = dias_a_devolver / dias_no_mes

        esp_saude_mes = m_saude * fracao
        esp_odonto_mes = m_odonto * fracao
        esp_saude += esp_saude_mes
        esp_odonto += esp_odonto_mes

        pbh_saude += esp_saude_mes * (s_saude / m_saude if m_saude > 0 else 0.0)
        pbh_odonto += esp_odonto_mes * (s_odonto / m_odonto if m_odonto > 0 else 0.0)

        mes_atual = proximo_mes(mes_atual)

    return pd.Series({
        'ESPERADO_SAUDE': round(esp_saude, 2),
        'PCT_PBH_SAUDE': (pbh_saude / esp_saude) if esp_saude > 0 else 0.0,
        'PCT_PBH_ODONTO': (pbh_odonto / esp_odonto) if esp_odonto > 0 else 0.0,
        'PERIODO_SUBSIDIADO_STR': formatar_periodo(meses_subsidiados),
    })


def checar_status(row) -> str:
    if pd.isna(row['ESPERADO_SAUDE']) or row['ESPERADO_SAUDE'] == 0:
        return '-'
    if abs(row['CRED_OPERADORA_SAUDE'] - row['ESPERADO_SAUDE']) <= 0.10:
        return 'OK'
    return f"ERRO: Op mandou R${row['CRED_OPERADORA_SAUDE']:.2f}, calc R${row['ESPERADO_SAUDE']:.2f}"


def main():
    print("Iniciando automação...")

    try:
        open(ARQUIVO_OFICIAL, 'r+').close()
    except PermissionError:
        print(f"\n[ERRO] '{ARQUIVO_OFICIAL}' está aberto no Excel. Feche e tente novamente.")
        sys.exit()

    df = pd.read_excel(ARQUIVO_OFICIAL, sheet_name='EMISSÃO', skiprows=1)
    df['CPF_CLEAN'] = df['CPF'].apply(clean_cpf)
    colunas_originais = df.columns.tolist()

    # Passivo Zetra
    print("Buscando passivo ZETRA...")
    try:
        arq = glob.glob('CONTROLE CONSOLIDADO PASSIVO GESTORA ZETRA*.xlsx')[0]
        df_z = pd.read_excel(arq, sheet_name='INADIMPLÊNCIA - PBH', skiprows=1)
        df_z['CPF_CLEAN'] = df_z['CPF'].apply(clean_cpf)
        df_z['SALDO FINAL'] = pd.to_numeric(df_z['SALDO FINAL'], errors='coerce').fillna(0)
        grp = df_z.groupby('CPF_CLEAN', as_index=False)['SALDO FINAL'].sum()
        df = pd.merge(df, grp, on='CPF_CLEAN', how='left')
        col = [c for c in colunas_originais if 'Nov/2021' in str(c)][0]
        df[col] = df['SALDO FINAL'].fillna(0)
        df.drop(columns=['SALDO FINAL'], inplace=True)
    except Exception as e:
        print(f"  Aviso: base Zetra não encontrada ou com erro: {e}")

    # Passivo Unimed
    print("Buscando passivo consolidado UNIMED...")
    try:
        arq = glob.glob('CONTROLE CONSOLIDADO DE INADIMPLENTES*.xlsx')[0]
        df_p = pd.read_excel(arq, sheet_name='UNIMED', skiprows=1)
        df_p['CPF_CLEAN'] = df_p['CPF'].apply(clean_cpf)
        grp = df_p.groupby('CPF_CLEAN', as_index=False)['SALDO CONSOLIDADO2'].sum()
        df = pd.merge(df, grp, on='CPF_CLEAN', how='left')
        col = [c for c in colunas_originais if '12/2021' in str(c)][0]
        df[col] = df['SALDO CONSOLIDADO2'].fillna(0)
        df.drop(columns=['SALDO CONSOLIDADO2'], inplace=True)
    except Exception:
        print("  Sem atualizações de passivo Unimed.")

    # Boletos em aberto
    print("Buscando boletos em aberto...")
    for arq in glob.glob('PBH - MNC BOLETO*.xlsx'):
        try:
            match = re.search(r'BOLETO (\d{2})\.(\d{4})', arq)
            if not match:
                continue
            mes, ano = match.groups()
            ref = f"ref. {mes}/{ano[-2:]}"
            col_destino = [c for c in colunas_originais if 'Boleto' in str(c) and ref in str(c)]
            if not col_destino:
                continue
            abas = pd.ExcelFile(arq).sheet_names
            aba = [a for a in abas if 'MNC' in a.upper()][0]
            df_bol = pd.read_excel(arq, sheet_name=aba)
            df_bol['CPF_CLEAN'] = df_bol['CPF Titular'].apply(clean_cpf)
            grp = df_bol.groupby('CPF_CLEAN', as_index=False)['TOTAL GERAL'].sum()
            df = pd.merge(df, grp, on='CPF_CLEAN', how='left')
            df[col_destino[0]] = df['TOTAL GERAL'].fillna(0)
            df.drop(columns=['TOTAL GERAL'], inplace=True)
        except Exception:
            pass

    # Enviados à folha
    print("Buscando enviados à folha...")
    for arq in glob.glob('PBH - ENVIADO À FOLHA*.xlsx'):
        try:
            match = re.search(r'FOLHA (\d{2})\.(\d{4})', arq)
            if not match:
                continue
            mes, ano = match.groups()
            ref = f"ref. {mes}/{ano[-2:]}"
            col_destino = [
                c for c in colunas_originais
                if ('Valores' in str(c) or 'plano' in str(c)) and ref in str(c)
            ]
            if not col_destino:
                continue
            abas = pd.ExcelFile(arq).sheet_names
            abas_validas = [a for a in abas if 'CONSOLIDADO' in a.upper() or 'FOLHA' in a.upper()]
            aba = abas_validas[0] if abas_validas else abas[0]
            df_f = pd.read_excel(arq, sheet_name=aba)
            df_f['CPF_CLEAN'] = df_f['CPF'].apply(clean_cpf)
            grp = df_f.groupby('CPF_CLEAN', as_index=False)['VALOR'].sum()
            df = pd.merge(df, grp, on='CPF_CLEAN', how='left')
            df[col_destino[0]] = df['VALOR'].fillna(0)
            df.drop(columns=['VALOR'], inplace=True)
        except Exception as e:
            print(f"  Aviso: erro ao ler folha {arq}: {e}")

    cols_debito = [
        c for c in colunas_originais
        if 'Débito passivo' in str(c) or 'Boleto' in str(c) or 'Valores de plano' in str(c)
    ]
    df[' TOTAL DÉBITOS'] = df[cols_debito].sum(axis=1)

    # Rateio mês a mês
    print("Carregando demonstrativos...")
    df['DATA_OBITO'] = pd.to_datetime(df['DATA ÓBITO'], errors='coerce', dayfirst=True)
    df['DATA_CANCELAMENTO'] = pd.to_datetime(df['DATA CANCELAMENTO'], errors='coerce', dayfirst=True)
    df_dem_pivot = carregar_demonstrativos()

    print("Calculando rateio mês a mês...")
    rateio = df.apply(lambda row: calcular_rateio(row, df_dem_pivot), axis=1)
    df['ESPERADO_SAUDE'] = rateio['ESPERADO_SAUDE']
    df['PERÍODO SUBSIDIADO STR'] = rateio['PERIODO_SUBSIDIADO_STR']

    # Créditos da operadora
    print("Buscando créditos da operadora...")
    try:
        arq_cred = glob.glob('CREDITOS*.xlsx')[0]

        df_cs = pd.read_excel(arq_cred, sheet_name='CRED_SUB_SAÚDE')
        df_cs['CPF_CLEAN'] = df_cs['CPF_Beneficiario'].apply(clean_cpf)
        cred_saude = df_cs.groupby('CPF_CLEAN', as_index=False)['Crédito abatido no Subsidio'].sum()
        cred_saude.rename(columns={'Crédito abatido no Subsidio': 'CRED_OPERADORA_SAUDE'}, inplace=True)

        df_co = pd.read_excel(arq_cred, sheet_name='CRED_SUB_ODONTO')
        df_co['CPF_CLEAN'] = df_co['CPF_Beneficiario'].apply(clean_cpf)
        cred_odonto = df_co.groupby('CPF_CLEAN', as_index=False)['Crédito abatido no Subsidio'].sum()
        cred_odonto.rename(columns={'Crédito abatido no Subsidio': 'CRED_OPERADORA_ODONTO'}, inplace=True)

        try:
            df_cg = pd.read_excel(arq_cred, sheet_name='CRED_SUB_GENÉRICO')
            df_cg['CPF_CLEAN'] = df_cg['CPF_Beneficiario'].apply(clean_cpf)
            cred_generico = df_cg.groupby('CPF_CLEAN', as_index=False)['Crédito abatido no Subsidio'].sum()
            cred_generico.rename(columns={'Crédito abatido no Subsidio': 'CRED_OPERADORA_GENERICO'}, inplace=True)
        except Exception:
            cred_generico = pd.DataFrame(columns=['CPF_CLEAN', 'CRED_OPERADORA_GENERICO'])

        df = pd.merge(df, cred_saude, on='CPF_CLEAN', how='left')
        df = pd.merge(df, cred_odonto, on='CPF_CLEAN', how='left')
        df = pd.merge(df, cred_generico, on='CPF_CLEAN', how='left')

    except IndexError:
        print("\n[ERRO] Arquivo de créditos da operadora não encontrado.")
        sys.exit()

    for col in ('CRED_OPERADORA_SAUDE', 'CRED_OPERADORA_ODONTO', 'CRED_OPERADORA_GENERICO'):
        df[col] = df[col].fillna(0)

    df['STATUS_AUDITORIA'] = df.apply(checar_status, axis=1)
    df['DEVOLVER_PBH_SAUDE'] = (df['CRED_OPERADORA_SAUDE'] * rateio['PCT_PBH_SAUDE']).round(2)
    df['CRÉDITO FALECIDO SAÚDE'] = (df['CRED_OPERADORA_SAUDE'] - df['DEVOLVER_PBH_SAUDE']).round(2)
    df['DEVOLVER_PBH_ODONTO'] = (df['CRED_OPERADORA_ODONTO'] * rateio['PCT_PBH_ODONTO']).round(2)
    df['CRÉDITO FALECIDO ODONTO'] = (df['CRED_OPERADORA_ODONTO'] - df['DEVOLVER_PBH_ODONTO']).round(2)

    # Escrita no Excel
    print("Salvando no Excel...")
    wb = load_workbook(ARQUIVO_OFICIAL)
    ws = wb['EMISSÃO']

    df['DATA_OBITO_STR'] = df['DATA_OBITO'].dt.strftime('%d/%m/%Y').fillna('')
    df['DATA_CANCELAMENTO_STR'] = df['DATA_CANCELAMENTO'].dt.strftime('%d/%m/%Y').fillna('')

    borda_fina = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    fundo_cabecalho = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)

    colunas_novas = [
        ('CRÉDITO ASSISTENCIAL',   'CRED_OPERADORA_SAUDE',    FORMATO_MOEDA),
        ('CRÉDITO ODONTO',         'CRED_OPERADORA_ODONTO',   FORMATO_MOEDA),
        ('CRÉDITO GENÉRICO',       'CRED_OPERADORA_GENERICO', FORMATO_MOEDA),
        ('ESPERADO SAÚDE',         'ESPERADO_SAUDE',          FORMATO_MOEDA),
        ('STATUS AUDITORIA',       'STATUS_AUDITORIA',        None),
        ('SUBSÍDIO PBH SAÚDE',     'DEVOLVER_PBH_SAUDE',      FORMATO_MOEDA),
        ('SUBSÍDIO PBH ODONTO',    'DEVOLVER_PBH_ODONTO',     FORMATO_MOEDA),
        ('CRÉDITO FALECIDO SAÚDE', 'CRÉDITO FALECIDO SAÚDE',  FORMATO_MOEDA),
        ('CRÉDITO FALECIDO ODONTO','CRÉDITO FALECIDO ODONTO', FORMATO_MOEDA),
        ('PERÍODO SUBSIDIADO',     'PERÍODO SUBSIDIADO STR',  None),
    ]

    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

    for index, row in df.iterrows():
        row_excel = index + 3
        if 'DATA ÓBITO' in headers:
            ws.cell(row=row_excel, column=headers.index('DATA ÓBITO') + 1, value=row['DATA_OBITO_STR'])
        if 'DATA CANCELAMENTO' in headers:
            ws.cell(row=row_excel, column=headers.index('DATA CANCELAMENTO') + 1, value=row['DATA_CANCELAMENTO_STR'])

        for col_nome in cols_debito + [' TOTAL DÉBITOS']:
            if col_nome in headers:
                cel = ws.cell(row=row_excel, column=headers.index(col_nome) + 1, value=row[col_nome])
                cel.number_format = FORMATO_MOEDA
                cel.border = borda_fina

    col_atual = ws.max_column + 1
    indices_novas = {}

    for nome_excel, col_df, fmt in colunas_novas:
        if nome_excel in headers:
            c_idx = headers.index(nome_excel) + 1
        else:
            c_idx = col_atual
            cab = ws.cell(row=2, column=c_idx, value=nome_excel)
            cab.fill = fundo_cabecalho
            cab.font = fonte_cabecalho
            cab.border = borda_fina
            cab.alignment = Alignment(horizontal='center', vertical='center')
            col_atual += 1
        indices_novas[nome_excel] = (c_idx, col_df, fmt)

    for index, row in df.iterrows():
        row_excel = index + 3
        for nome_excel, (c_idx, col_df, fmt) in indices_novas.items():
            if col_df and col_df in row.index and pd.notna(row[col_df]):
                cel = ws.cell(row=row_excel, column=c_idx, value=row[col_df])
                if fmt:
                    cel.number_format = fmt
                cel.border = borda_fina

    wb.save(ARQUIVO_OFICIAL)
    print("\n[OK] Planilha atualizada com sucesso.")


if __name__ == '__main__':
    main()
